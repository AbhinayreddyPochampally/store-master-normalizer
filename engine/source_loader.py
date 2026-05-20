"""Load a brand's source data sheet into a list of row dicts.

Caller supplies the sheet name and the (1-indexed) header row explicitly --
no auto-detection.  For sheets that contain duplicate header names (e.g.
PF Active Stores has 'Contact No', 'Mail Id', 'Mobile No' and 'ASP Code'
each twice), every occurrence is preserved with its column index so that
mapper.py can resolve them by adjacency later.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple

import openpyxl
from pyxlsb import open_workbook as open_xlsb


@dataclass
class SourceSheet:
    """In-memory representation of a brand's data sheet.

    Attributes
    ----------
    columns: list of (col_index, header_name) in original order.  Header names
        are kept verbatim -- trailing spaces and special characters are NOT
        stripped, because the reference sheet may reference them that way.
    rows: list of {col_index: cell_value} dicts.
    name_to_indices: header_name -> [col_index, ...] for duplicate lookup.
    """

    columns: List[Tuple[int, str]]
    rows: List[Dict[int, Any]]
    name_to_indices: Dict[str, List[int]] = field(default_factory=dict)

    def indices_for(self, name: str) -> List[int]:
        return self.name_to_indices.get(name, [])

    def column_index(self, name: str) -> int:
        """Return the single column index for a name; raise if duplicated."""
        idxs = self.indices_for(name)
        if not idxs:
            raise KeyError(f"No column named {name!r}")
        if len(idxs) > 1:
            raise KeyError(
                f"Column {name!r} is duplicated at indices {idxs}; caller "
                f"must resolve by adjacency."
            )
        return idxs[0]


def _build_name_index(columns: List[Tuple[int, str]]) -> Dict[str, List[int]]:
    out: Dict[str, List[int]] = {}
    for idx, name in columns:
        if name is None:
            continue
        out.setdefault(name, []).append(idx)
    return out


def _is_blank_row(values) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


# -- xlsx loader ---------------------------------------------------------

def _load_xlsx(path: str, sheet_name: str, header_row: int) -> SourceSheet:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header_values = None
        data_rows: List[Dict[int, Any]] = []
        # openpyxl iterates from row 1.  header_row is 1-indexed (matches CLI).
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < header_row:
                continue
            if i == header_row:
                header_values = list(row)
                continue
            # Skip entirely blank rows (common at sheet tail).
            if _is_blank_row(row):
                continue
            data_rows.append({j: row[j] if j < len(row) else None
                              for j in range(len(header_values))})
        if header_values is None:
            raise ValueError(f"Sheet {sheet_name!r} has no row {header_row}.")
        columns = [(j, header_values[j]) for j in range(len(header_values))]
        return SourceSheet(columns=columns, rows=data_rows,
                           name_to_indices=_build_name_index(columns))
    finally:
        wb.close()


# -- xlsb loader ---------------------------------------------------------

def _load_xlsb(path: str, sheet_name: str, header_row: int) -> SourceSheet:
    header_values = None
    data_rows: List[Dict[int, Any]] = []
    with open_xlsb(path) as wb:
        with wb.get_sheet(sheet_name) as sheet:
            for i, row in enumerate(sheet.rows(), start=1):
                vals = [c.v for c in row]
                if i < header_row:
                    continue
                if i == header_row:
                    header_values = vals
                    continue
                if _is_blank_row(vals):
                    continue
                # Pad / trim to header width.
                width = len(header_values)
                data_rows.append({j: vals[j] if j < len(vals) else None
                                  for j in range(width)})
    if header_values is None:
        raise ValueError(f"Sheet {sheet_name!r} has no row {header_row}.")
    columns = [(j, header_values[j]) for j in range(len(header_values))]
    return SourceSheet(columns=columns, rows=data_rows,
                       name_to_indices=_build_name_index(columns))


# -- Public API ----------------------------------------------------------

def load_source(path: str, sheet_name: str, header_row: int) -> SourceSheet:
    """Open the given workbook and load one data sheet.

    Parameters
    ----------
    path: workbook path (.xlsx or .xlsb).
    sheet_name: exact sheet name -- no fuzzy match.
    header_row: 1-indexed row number where headers live.  Data starts at
        header_row + 1.
    """
    lower = path.lower()
    if lower.endswith(".xlsb"):
        return _load_xlsb(path, sheet_name, header_row)
    return _load_xlsx(path, sheet_name, header_row)


# -- Sheet-list inspection (v0.5.2) --------------------------------------

def list_sheets(path: str) -> List[str]:
    """Return the sheet names in a workbook without loading any data.

    Used by the v0.5.2 per-row sheet picker so the UI can surface
    'Available: ...' diagnostics when the operator's typed sheet name
    doesn't match.  Runs in <100 ms for typical inputs because both
    backends only read the directory structure of the .xlsx zip /
    workbook stream, not the cells.
    """
    if path.lower().endswith(".xlsb"):
        with open_xlsb(path) as wb:
            return list(wb.sheets)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def sheet_has_data(path: str, sheet_name: str, header_row: int) -> bool:
    """Return True if ``sheet_name`` has at least one non-header data row.

    Cheap probe used by the per-row sheet picker's onBlur validation.
    Stops at the first non-blank data row encountered, so even large
    sheets complete in well under a second.
    """
    if path.lower().endswith(".xlsb"):
        with open_xlsb(path) as wb:
            if sheet_name not in list(wb.sheets):
                return False
            with wb.get_sheet(sheet_name) as sheet:
                for i, row in enumerate(sheet.rows(), start=1):
                    if i <= header_row:
                        continue
                    vals = [c.v for c in row]
                    if not _is_blank_row(vals):
                        return True
        return False
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= header_row:
                continue
            if not _is_blank_row(row):
                return True
        return False
    finally:
        wb.close()
