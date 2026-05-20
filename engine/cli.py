"""Command-line entry point for the store-master normalizer.

Example
-------
::

    python -m engine.cli \
        --source "inputs/Pantaloons_Apr_2026.xlsx" \
        --master "inputs/Backend_Data_-_Store_Master.xlsx" \
        --brand-name Pantaloons \
        --scope-column brand \
        --scope-value Pantaloons \
        --sheet Sheet1 \
        --header-row 1 \
        --out-dir outputs

Writes:
    <out-dir>/<brand>_Updated_Master_<YYYY-MM-DD>.xlsx
    <out-dir>/<brand>_Change_Report_<YYYY-MM-DD>.xlsx
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os
import sys
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

# Allow running either as ``python -m engine.cli`` or ``python engine/cli.py``.
if __package__ in (None, ""):
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from engine.reference_reader import get_mapping_for_brand, Rule, COLUMN
    from engine.source_loader import load_source
    from engine.mapper import map_all, _coerce_int
    from engine.brand_overrides import get as get_overrides
    from engine.reconciler import reconcile, _canonical, apply_inactivation_pass
    from engine.dates import parse_iso_date
    from engine.brands import load_brands
else:
    from .reference_reader import get_mapping_for_brand, Rule, COLUMN
    from .source_loader import load_source
    from .mapper import map_all, _coerce_int
    from .brand_overrides import get as get_overrides
    from .reconciler import reconcile, _canonical, apply_inactivation_pass
    from .dates import parse_iso_date
    from .brands import load_brands


# Fields normalised on every write (regardless of which engine path
# touched the row).  v0.4.1 A2/A3: applies to legacy rows too.
_ISO_DATE_FIELDS = ("Store Opening Date",)
_INT_FIELDS = ("Pincode",)


REFERENCE_SHEET_GUESSES = [
    "{brand} Reference",
    "{brand}-Reference",
    "{brand}Reference",
    "PF Reference",
    "TCNS-Reference",
    "Tasva Reference",
    "Pantaloons Reference",
]


def _detect_reference_sheet(path: str, brand: str) -> str:
    if path.lower().endswith(".xlsb"):
        from pyxlsb import open_workbook as open_xlsb
        with open_xlsb(path) as wb:
            sheets = list(wb.sheets)
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    candidates = [g.format(brand=brand) for g in REFERENCE_SHEET_GUESSES]
    for c in candidates:
        if c in sheets:
            return c
    for s in sheets:
        if "reference" in s.lower():
            return s
    raise SystemExit(f"Could not find a reference sheet in {path!r}. "
                     f"Sheets present: {sheets}")


def _load_master(path: str,
                 sheet: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load the backend-master workbook into row dicts.

    v0.5.2: ``sheet`` is now overridable.  Default is ``Sheet1`` for backward
    compatibility with pre-v0.5 backend masters; the FastAPI layer currently
    passes ``Backend Updated Data`` from ``_backend.default_sheet`` in
    brands.json (or whatever the operator typed into the UI).

    Raises ValueError with the list of available sheets if ``sheet`` is not
    in the workbook -- the web layer surfaces this as an inline error.
    """
    sheet = sheet or "Sheet1"
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet!r} not found in backend master. "
                f"Available: {wb.sheetnames!r}"
            )
        ws = wb[sheet]
        header_values: List[str] = []
        rows: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                header_values = [str(h) if h is not None else f"_col{j}"
                                 for j, h in enumerate(row)]
                continue
            if all(v is None for v in row):
                continue
            rows.append({header_values[j]: row[j] if j < len(row) else None
                         for j in range(len(header_values))})
        return rows, header_values
    finally:
        wb.close()


def _apply_field_overrides(rules: List[Rule], overrides) -> List[Rule]:
    if not overrides.field_overrides:
        return rules
    out = list(rules)
    by_target = {r.target_field: i for i, r in enumerate(out)}
    for target, src_type, src_value in overrides.field_overrides:
        replacement = Rule(target_field=target,
                           source_type=src_type,
                           source_value=src_value,
                           notes="(applied via brand_overrides)")
        if target in by_target:
            out[by_target[target]] = replacement
        else:
            out.append(replacement)
            by_target[target] = len(out) - 1
    return out


def _resolve_col_ref(ref, sheet) -> Optional[Tuple[int, str]]:
    """Resolve a divergence-pair column reference to (col_index, label).

    `ref` may be an int (1-indexed column position) or a str (column name).
    Returns None if not present in the sheet.
    """
    if isinstance(ref, int):
        idx_0 = ref - 1
        if 0 <= idx_0 < len(sheet.columns):
            return idx_0, f"col {ref} ({sheet.columns[idx_0][1]!r})"
        return None
    # String: look up by name.  Duplicates -> leftmost (operator can refine
    # by passing an int instead).
    idxs = sheet.indices_for(ref)
    if not idxs:
        return None
    return idxs[0], f"{ref!r}"


def _check_divergences(sheet, mapped_rows, overrides, warnings: List[str],
                       brand_label: str = "") -> None:
    """Compare each row's value between two columns flagged as having
    historically drifted apart.

    v0.4.1 C4: warnings are emitted in the multi-line format the UI
    parses into an expandable block:

        WC36 -- TCNS -- ASP Code mismatch
          Master col 1: 6812
          Master col 55: 3442
          Action: review and choose canonical value

    The leading line is the "summary" the UI shows collapsed; the
    indented lines are the expand-on-click detail.
    """
    if not overrides.divergence_warnings:
        return
    for col_a, col_b in overrides.divergence_warnings:
        ra = _resolve_col_ref(col_a, sheet)
        rb = _resolve_col_ref(col_b, sheet)
        if ra is None or rb is None:
            warnings.append(
                f"Divergence check skipped: {col_a!r} or {col_b!r} not in source sheet."
            )
            continue
        ia, label_a = ra
        ib, label_b = rb
        # Derive a short "field" name from label_a, stripping the
        # "col N (..." prefix when possible.
        def _short(lbl: str) -> str:
            if "(" in lbl and lbl.endswith(")"):
                return lbl[lbl.index("(") + 2 : -2]  # strip "(' ... ')"
            return lbl
        field_short = _short(label_a)
        for r_no, (row, mapped) in enumerate(zip(sheet.rows, mapped_rows), start=1):
            va, vb = row.get(ia), row.get(ib)
            if _canonical(va) is None or _canonical(vb) is None:
                continue
            if _canonical(va) != _canonical(vb):
                sid = (mapped.get("Store Id") if mapped else None) or f"row {r_no}"
                head = (
                    f"{sid} -- {brand_label or 'source'} -- "
                    f"{field_short} mismatch"
                ) if brand_label else (
                    f"{sid} -- {field_short} mismatch"
                )
                detail = (
                    f"\n  Master {label_a}: {va!r}"
                    f"\n  Master {label_b}: {vb!r}"
                    f"\n  Action: review and choose canonical value"
                )
                warnings.append(head + detail)


def _write_updated_master(path: str, rows: List[Dict[str, Any]],
                          field_order: List[str]) -> None:
    # v0.4.1 A2 + A3: once-through normalisation of ISO-date and
    # int fields, so rows that weren't touched by the source this run
    # still come out in YYYY-MM-DD / int instead of carrying their
    # legacy master format forward.  We mutate in place; the rows are
    # the engine's working copies, not the caller's master.
    for r in rows:
        for f in _ISO_DATE_FIELDS:
            if f in r:
                r[f] = parse_iso_date(r.get(f))
        for f in _INT_FIELDS:
            if f in r:
                r[f] = _coerce_int(r.get(f))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"   # v0.4.1 A5: sheet named "Master", not "Sheet1"
    ws.append(field_order)
    for r in rows:
        ws.append([r.get(f) for f in field_order])
    wb.save(path)


# v0.4.1 A5 / B2: map engine status codes to the user-facing
# Classification labels the change report uses.
_CLASSIFICATION_LABELS = {
    "UPDATED":      "Refresh",
    "CODE_CHANGED": "Migrated",
    "REACTIVATED":  "Reactivated",
    "NEW":          "New",
    "CLOSED":       "Inactivated-missing",
    "INACTIVATED_BAD_EMAIL": "Inactivated-bad-email",
    "ORPHAN":       "Orphan",
}


def _write_change_report(path: str, summary: Dict[str, int],
                         changes, source_total: int,
                         matched: int, unmatched: int, orphan: int,
                         warnings: List[str]) -> None:
    wb = openpyxl.Workbook()
    sw = wb.active
    sw.title = "Summary"
    sw.append(["Metric", "Count"])
    sw.append(["Source rows", source_total])
    sw.append(["Matched to master (in-scope)", matched])
    sw.append(["Unmatched (NEW)", unmatched])
    sw.append(["Orphan (out-of-scope match)", orphan])
    sw.append([])
    sw.append(["Classification", "Count"])
    for k in ("NEW", "UPDATED", "CODE_CHANGED", "REACTIVATED",
              "CLOSED", "INACTIVATED_BAD_EMAIL", "ORPHAN"):
        if summary.get(k, 0):
            sw.append([_CLASSIFICATION_LABELS.get(k, k), summary[k]])
    sw.append([])
    sw.append(["Warnings", len(warnings)])
    for w in warnings:
        sw.append(["", w])

    # v0.4.1 A5: per-spec column names + ordering.
    cw = wb.create_sheet("Changes")
    cw.append(["Store Id", "Classification", "Field",
               "Before", "After", "Remark"])
    for c in changes:
        cw.append([
            c.store_id,
            _CLASSIFICATION_LABELS.get(c.status, c.status),
            c.field_changed,
            c.old_value,
            c.new_value,
            c.notes,
        ])
    wb.save(path)


# v0.4.1 A5: produce filenames like "Pantaloons_Updated_Master_May-2026.xlsx".
# month_label comes from the Setup form (web) or --month (CLI).  If absent
# we fall back to today's "Mon-YYYY" so we never hand back a hashed name.
_MONTH_NAMES = (
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
)


def _format_month_label(raw: Optional[str]) -> str:
    if raw:
        return raw.strip().replace("/", "-").replace(" ", "-")
    today = _dt.date.today()
    return f"{_MONTH_NAMES[today.month - 1]}-{today.year}"


def make_output_paths(out_dir: str, brand_slug: str,
                      month_label: Optional[str]) -> Tuple[str, str]:
    label = _format_month_label(month_label)
    updated = os.path.join(out_dir, f"{brand_slug}_Updated_Master_{label}.xlsx")
    report = os.path.join(out_dir, f"{brand_slug}_Change_Report_{label}.xlsx")
    return updated, report


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Normalize a brand's store-master export into the backend master schema."
    )
    p.add_argument("--source", required=True)
    p.add_argument("--master", required=True)
    p.add_argument("--brand-name", required=True)
    p.add_argument("--scope-column", required=True)
    p.add_argument("--scope-value", required=True)
    p.add_argument("--sheet", required=True)
    p.add_argument("--header-row", required=True, type=int)
    p.add_argument("--out-dir", default="outputs")
    p.add_argument("--reference-sheet", default=None)
    # v0.4.1 A5: month label for output filenames, e.g. "May-2026".
    p.add_argument("--month", default=None,
                   help="Month label baked into output filenames "
                        "(e.g. 'May-2026').  Defaults to current month.")

    args = p.parse_args(argv)
    os.makedirs(args.out_dir, exist_ok=True)

    rules = get_mapping_for_brand(args.brand_name)
    overrides = get_overrides(args.brand_name)

    sheet = load_source(args.source, args.sheet, args.header_row)
    warnings: List[str] = []
    mapped, map_warnings = map_all(rules, sheet)
    warnings.extend(map_warnings)

    # Divergence checks run after mapping so we can label warnings with Store Id.
    _check_divergences(sheet, mapped, overrides, warnings)

    master_rows, master_field_order = _load_master(args.master)
    result = reconcile(
        rules=rules,
        mapped_rows=mapped,
        master_rows=master_rows,
        master_field_order=master_field_order,
        scope_column=args.scope_column,
        scope_value=args.scope_value,
        month_label=args.month,
    )

    # v0.4.1 B2: Section 6 inactivation pass runs AFTER the cascade.
    brand_cfg = load_brands().get(args.brand_name, {})
    apply_inactivation_pass(
        result=result,
        brand_label=brand_cfg.get("label", args.brand_name),
        brand_label_short=args.brand_name,
        month_label=args.month,
    )
    warnings.extend(result.warnings)

    brand_slug = args.brand_name.replace(" ", "_")
    updated_path, report_path = make_output_paths(
        args.out_dir, brand_slug, args.month,
    )

    _write_updated_master(updated_path, result.updated_master, master_field_order)
    _write_change_report(report_path,
                         summary=result.summary,
                         changes=result.changes,
                         source_total=len(mapped),
                         matched=result.matched_source_count,
                         unmatched=result.unmatched_source_count,
                         orphan=result.orphan_source_count,
                         warnings=warnings)

    print(f"Wrote {updated_path}")
    print(f"Wrote {report_path}")
    print()
    print("Summary:")
    for k in ("NEW", "UPDATED", "CODE_CHANGED", "REACTIVATED",
              "CLOSED", "INACTIVATED_BAD_EMAIL", "ORPHAN"):
        if result.summary.get(k, 0):
            print(f"  {k:24} {result.summary[k]}")
    print()
    print(f"Source rows:               {len(mapped)}")
    print(f"  Matched to master:       {result.matched_source_count}")
    print(f"  Unmatched (NEW):         {result.unmatched_source_count}")
    print(f"  Orphan (out-of-scope):   {result.orphan_source_count}")
    if warnings:
        print()
        print(f"Warnings ({len(warnings)}):")
        for w in warnings[:30]:
            print(f"  - {w}")
        if len(warnings) > 30:
            print(f"  ... and {len(warnings)-30} more (see Summary sheet)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
